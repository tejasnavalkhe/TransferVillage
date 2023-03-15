from botocore.exceptions import ClientError
import boto3
from TransferVillage.config import Config
from docx2pdf import convert
from pdf2docx import Converter
from fpdf import FPDF
from docx import Document
import img2pdf
from PIL import Image
from openpyxl import load_workbook

class BOTO_S3:
	def upload_file(self, original_fileobj, filename, folder_name=None):
		"""Upload files to S3 object

		Args:
			original_fileobj (str): The string containing the original file name
			filename (str): The string containing the unique filename
			folder_name (str, optional): The string containing the folder name. Defaults to None.

		Returns:
			Boolean: Return True if the file was uploaded successfully else False
		"""
		try:
			s3 = boto3.resource(Config.AWS_SERVICE_NAME)
			if folder_name:
				s3.Bucket(Config.AWS_BUCKET_NAME).upload_fileobj(original_fileobj, f"{folder_name}/{filename}")
			else:
				s3.Bucket(Config.AWS_BUCKET_NAME).upload_fileobj(original_fileobj, f"{filename}")
			return True
		except ClientError as e:
			print(e)
			return False

	def delete_file(self, filename, folder_name=None):
		"""Delete file which is requested

		Args:
			filename (str): The string containing file name
			folder_name (str, optional): The string containing folder name. Defaults to None.

		Returns:
			Boolean: Return True if file was successfully deleted else False.
		"""
		try:
			s3 = boto3.resource(Config.AWS_SERVICE_NAME)
			if folder_name:
				s3.Bucket(Config.AWS_BUCKET_NAME).Object(f"{folder_name}/{filename}").delete()
			else:
				s3.Bucket(Config.AWS_BUCKET_NAME).Object(filename).delete()
			return True
		except ClientError as e:
			print(e)
			return False
	
	def delete_folder(self, folder_name):
		"""Delete an empty folder

		Args:
			folder_name (str): A string containing folder name

		Returns:
			Boolean: Return True if deleted else False
		"""
		try:
			s3 = boto3.resource(Config.AWS_SERVICE_NAME)
			bucket = s3.Bucket(Config.AWS_BUCKET_NAME)
			bucket.objects.filter(Prefix=f"{folder_name}/").delete()
			return True
		except ClientError as e:
			print(e)
			return False

	def create_presigned_url(self, filename, expiration=3600):
		"""Generate a presigned URL to share an S3 object

		Args:
			filename (str): A string containing file name
			expiration (int): Expiration time. After this time the access to this S3 object will be restricted.

		Returns:
			str: Presigned URL as string. If error, returns None.
		"""

		# Generate a presigned URL for the S3 object
		s3_client = boto3.client(Config.AWS_SERVICE_NAME)
		try:
			response = s3_client.generate_presigned_url('get_object',
														Params={'Bucket': Config.AWS_BUCKET_NAME,
																'Key': filename},
														ExpiresIn=expiration)
		except ClientError as e:
			print(e)
			return None

		# The response contains the presigned URL
		return response

class FileHandler:
	def allowed_file(self, filename):
		return '.' in filename and filename.rsplit('.', 1)[1].lower() in Config.ALLOWED_EXTENSIONS

	def WORD_TO_PDF(self, word_file):
		convert(word_file)
	
	def TXT_TO_PDF(self, txt_obj):
		# save FPDF() class into a variable pdf
		pdf = FPDF()

		# Add a page
		pdf.add_page()

		# set style and size of font that you want in the pdf
		pdf.set_font("Arial", size = 15)

		# insert the texts in pdf
		for x in txt_obj:
			pdf.cell(200, 10, txt = x, ln = 1, align = 'J')
		return pdf
	
	def IMAGE_TO_PDF(self, image_file, pdf_file):
		# opening image
		image = Image.open(image_file)
		# converting into chunks using img2pdf
		pdf_bytes = img2pdf.convert(image.filename)
		# opening or creating pdf file
		pdf = open(pdf_file, "wb")
		# writing pdf files with chunks
		pdf.write(pdf_bytes)
		# closing image file
		image.close()
		# closing pdf file
		pdf.close()

	def EXCEL_TO_PDF(self, excel_file):
		workbook = load_workbook(excel_file)
		sheet = workbook.active
		pdf = FPDF()
		pdf.add_page()
		for row in sheet.iter_rows():
			for cell in row:
				pdf.set_font('Arial', 'B', 10)
				pdf.cell(200, 10, str(cell.value))
		return pdf

	def PDF_TO_WORD(self, pdf_file):
		cv = Converter(pdf_file)
		cv.convert(pdf_file.rsplit('.', 1)[0]+'.docx')

	def TXT_TO_WORD(self, txt_obj):
		# create empty document
		doc = Document()
		# read text from file
		line = txt_obj.read()
		# add text to document
		doc.add_paragraph(line)
		return doc
